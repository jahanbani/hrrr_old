import os
from IPython import embed
import random
import sys
import time

import gams
import gamstransfer as gt
import ipdb
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

import folium


# import vincent
# import geopandas as gpd
from branca.element import Template, MacroElement

from openpyxl import load_workbook


def get_load(sdt, edt, area=321, directory=r"./load/"):
    """get the load from FERC 714 for the area
        COMBINE the following
        Carolina Duke Energy Carolinas: 157
        Progress Energy: 233
        south carolina energy and gas: 250
        south carolina public: 251

        "Duke Energy Corp: 134" is there load data for this
    areas = {"ISONE": 185, "NYISO": 211, "PJM": 230}
    south = [157, 233, 250, 251]
    MISO is 321
    """

    fn = directory + "Part 3 Schedule 2 - Planning Area Hourly Demand.csv"

    loadraw = pd.read_csv(fn)

    years = [sdt.year] + [edt.year]

    loadraw = loadraw[
        ["respondent_id", "report_yr", "plan_date"]
        + ["hour" + str(i).zfill(2) for i in range(1, 25)]
    ]
    load = loadraw.loc[
        (loadraw["respondent_id"] == area) & (loadraw["report_yr"].isin(years))
    ]

    loadn = load.drop(columns=["report_yr", "respondent_id", "plan_date"])
    loadn = loadn.stack().reset_index(drop=True)
    # loadn = loadn / 1000
    loadn.index = pd.date_range(
        start=str(sdt.year) + "-01-01 00:00",
        end=str(edt.year) + "-01-01 00:00",
        freq="1H",
        inclusive="left",
    )
    mask = (loadn.index >= sdt) & (loadn.index <= edt)
    loadn = loadn.loc[mask]
    return loadn


def build_inv_map(dfvars, Line_map=0, Gen_map=0, DR_map=0):
    print("MAPPING RESULTS")

    # import data
    filename = r".\Visualizations\MTEP_CODY.xlsx"
    data = pd.read_excel(filename, sheet_name=None)
    Summary = data["Summary"]
    Lines = data["Lines_ALL_EI"]
    Circles = data["Circles_USE_THIS"]
    Circles = Circles.loc[Circles["Number"] >
                          150][["Number", "Name", "Lon", "Lat"]]

    # parameters
    D = 0.97
    ROTATE = 90
    icon_size = 2

    # Create MAP
    iconURL = "https://image.flaticon.com/icons/svg/18/18554.svg"
    iconURL2 = "https://image.flaticon.com/icons/svg/148/148766.svg"
    iconHVDC = "http://teezeria.com/images/designs/558-Rimg_23.png"
    arrowURL = "http://www.iconsdb.com/icons/preview/blue/arrow-33-xxl.png"

    # EE - triangle
    # DR -
    iconEE = "http://teezeria.com/images/designs/558-Rimg_23.png"
    iconDR = "http://teezeria.com/images/designs/558-Rimg_23.png"

    wline = Summary.Parameter[1]
    warrow = Summary.Parameter[2]

    miso_map = folium.Map(
        location=[44, -78],
        zoom_start=5,
        min_zoom=1,
        max_zoom=20,
        tiles="OpenStreetMap",
        control_scale=True,
    )
    ######################### MISO AREA MAPPING ###############################
    """
    #Add ISO's
    print("MAPPING MISO SHAPE")
    ISO_area = gpd.GeoDataFrame.from_file(r'.\Visualizations\From Rajaz\Independent_System_Operators\Independent_System_Operators.shp')
    #print(len(ISO_area))

    ISO_area.drop(ISO_area[ISO_area['NAME'] != 'MIDCONTINENT INDEPENDENT TRANSMISSION SYSTEM OPERATOR, INC..'].index, inplace = True)
    #print(len(ISO_area))

    #number_of_colors = len(ISO_area)
    #print('number of colors = {}'.format(number_of_colors))

    #color =  ['#9dd1a5']
    color = ['#FF7F24']
    #print(color)

    ISO_area['RGB'] = color

    ISO_regions = folium.FeatureGroup(name='ISO territories')

    gjson = ISO_area[['NAME','geometry', 'ADDRESS', 'CITY','STATE','ZIP','COUNTRY','SOURCE','WEBSITE','RGB']].to_json()

    ISO_regions.add_child(folium.GeoJson(gjson,
                   name='Balancing Areas',
                   style_function=lambda feature: {'fillColor': feature['properties']['RGB'],
                                                   'color': 'black',
                                                   'weight': 3,
                                                   'dashArray': '5, 5',
                                                   'fillOpacity':0.07
                                                  },
                   highlight_function=lambda x: {'weight':3,
                                                 'fillColor': 'gold',
                                                 'color':'black',
                                                 'fillOpacity':1
                                                },
                   tooltip=folium.features.GeoJsonTooltip(
                       fields=['NAME', 'ADDRESS', 'CITY','STATE','ZIP','COUNTRY','SOURCE','WEBSITE'],
                       aliases=['Electric Retail Service Territory:', 'ADDRESS', 'CITY:','STATE:','ZIP:','COUNTRY:','SOURCE:','WEBSITE:'])))


    miso_map.add_child(ISO_regions)
    """
    ############### Mapping Invested Transmission Lines #######################
    if Line_map != 0 and "pv_LineInv" in dfvars.keys():
        print("MAPPING INVESTED TRANSMISSION LINES")
        Edges = folium.FeatureGroup(name="Flow")
        df = dfvars["pv_LineInv"].sort_values(by="level", ascending=False)
        df = df.loc[df["level"] > 100][["From", "To", "Lid", "Year", "level"]]
        # print(df.dtypes)
        convert_dict = {"From": object, "To": object, "Lid": object}
        df = df.astype(convert_dict)
        # print(df.dtypes)

        LineINVtot = pd.DataFrame()
        LineINVtot = df.groupby(["From", "To", "Lid"])[
            "level"].sum().reset_index()
        LineINVtot = (
            LineINVtot[LineINVtot["level"] >= 1]
            .sort_values(by="level", ascending=False)
            .reset_index(drop=True)
        )
        convert_dict = {"From": int, "To": int, "Lid": int}
        LineINVtot = LineINVtot.astype(convert_dict)
        LineINVtot = pd.concat(
            [
                LineINVtot,
                pd.DataFrame(
                    columns=[
                        "FromBus_name",
                        "ToBus_name",
                        "FromLon",
                        "FromLat",
                        "ToLon",
                        "ToLat",
                    ]
                ),
            ]
        )
        for i in np.arange(0, np.shape(LineINVtot)[0]):
            LineINVtot_slice = LineINVtot.iloc[i]
            Lid = LineINVtot_slice["Lid"]

            LineINVtot.at[i, "FromBus_name"] = Lines.loc[Lines["Line"] == Lid][
                "FromBus_name"
            ].values[0]
            LineINVtot.at[i, "ToBus_name"] = Lines.loc[Lines["Line"] == Lid][
                "ToBus_name"
            ].values[0]
            LineINVtot.at[i, "FromLon"] = Lines.loc[Lines["Line"] == Lid][
                "FromLon"
            ].values[0]
            LineINVtot.at[i, "FromLat"] = Lines.loc[Lines["Line"] == Lid][
                "FromLat"
            ].values[0]
            LineINVtot.at[i, "ToLon"] = Lines.loc[Lines["Line"] == Lid]["ToLon"].values[
                0
            ]
            LineINVtot.at[i, "ToLat"] = Lines.loc[Lines["Line"] == Lid]["ToLat"].values[
                0
            ]

        convert_dict = {
            "FromBus_name": str,
            "ToBus_name": str,
            "FromLon": float,
            "FromLat": float,
            "ToLon": float,
            "ToLat": float,
            "From": int,
            "To": int,
            "Lid": int,
        }
        LineINVtot = LineINVtot.astype(convert_dict)

        line_colors = ["black", "palegreen", "LimeGreen", "Green", "DarkGreen"]

        for i in np.arange(0, np.shape(LineINVtot)[0]):
            line_num = int(LineINVtot.Lid[i])

            CAP = LineINVtot.level[i]
            if np.greater(CAP, 0):
                popup = (
                    "Line ("
                    + str(LineINVtot.Lid[i])
                    + "): "
                    + str(LineINVtot.From[i])
                    + ">>>"
                    + str(LineINVtot.To[i])
                    + " | Capacity: "
                    + str(np.abs(round(CAP, 2)))
                    + " MW"
                )
                tooltip = (
                    "Line ("
                    + str(LineINVtot.Lid[i])
                    + "): "
                    + str(LineINVtot.From[i])
                    + ">>>"
                    + str(LineINVtot.To[i])
                    + " | Capacity: "
                    + str(np.abs(round(CAP, 2)))
                    + " MW"
                )

            weight = LineINVtot["level"][i] / 1000
            weight_l = 10
            if np.less(weight, 1) and np.greater(weight, 0):
                l_color = line_colors[0]
                weight_l = 2.5
            elif np.less(weight, 2) and np.greater(weight, 1):
                l_color = line_colors[1]
                weight_l = 5
            elif np.less(weight, 3) and np.greater(weight, 2):
                l_color = line_colors[2]
                weight_l = 7.5
            elif np.less(weight, 4) and np.greater(weight, 3):
                l_color = line_colors[3]
                weight_l = 10
            elif np.less(weight, 10) and np.greater(weight, 4):
                l_color = line_colors[4]
                weight_l = 12.5
            DD1 = np.abs(LineINVtot.FromLon[i] - LineINVtot.ToLon[i]) / D
            DD2 = np.abs(LineINVtot.FromLat[i] - LineINVtot.ToLat[i]) / D
            DD = int(np.max([DD1, DD2]))
            if DD == 0:
                DD = 1
            ArrowLong = np.round(
                np.linspace(LineINVtot.FromLon[i],
                            LineINVtot.ToLon[i], DD + 2),
                decimals=4,
            )
            ArrowLat = np.round(
                np.linspace(LineINVtot.FromLat[i],
                            LineINVtot.ToLat[i], DD + 2),
                decimals=4,
            )
            Edges.add_child(
                folium.PolyLine(
                    [[ArrowLat[0], ArrowLong[0]], [ArrowLat[1], ArrowLong[1]]],
                    weight=weight_l,
                    popup=popup,
                    tooltip=tooltip,
                    color=l_color,
                )
            )
            for j in np.arange(1, DD + 1):
                Edges.add_child(
                    folium.PolyLine(
                        [
                            [ArrowLat[j], ArrowLong[j]],
                            [ArrowLat[j + 1], ArrowLong[j + 1]],
                        ],
                        weight=weight_l,
                        popup=popup,
                        tooltip=tooltip,
                        color=l_color,
                    )
                )

        miso_map.add_child(Edges)
    ######################## Mapping Nodes #########################################
    Circles = Circles.loc[
        Circles["Number"].isin(
            LineINVtot["From"].to_list() + LineINVtot["To"].to_list()
        )
    ].reset_index()
    Nodes = folium.FeatureGroup(name="Nodes")
    print("MAPPING BUSES")
    LAT = Circles.Lat
    LON = Circles.Lon
    names = list(map(str, Circles.Number.values))  # + ' ' + Circles.Name
    for i in np.arange(0, LAT.size):
        # iframe = folium.IFrame(names[i], width=30, height=45)
        # popup = folium.Popup(iframe, max_width=3000)
        popup = names[i]
        Nodes.add_child(
            folium.CircleMarker(
                [LAT[i], LON[i]],
                popup=popup,
                tooltip=popup,
                icon=folium.features.CustomIcon(
                    icon_image=iconURL2, icon_size=(icon_size, icon_size)
                ),
                color="black",
            )
        )

    miso_map.add_child(Nodes)
    ##################### Mapping Invested Generation################################
    if Gen_map != 0 and "pv_GenInvYear" in dfvars.keys():
        print("MAPPING INVESTED GENERATION")
        df = dfvars["pv_GenInvYear"].sort_values(by="level", ascending=False)
        df = df.loc[df["level"] > 0.001][[
            "Bus", "Tech", "Gid", "Year", "level"]]
        # print(df.dtypes)
        convert_dict = {"Bus": object, "Tech": object,
                        "Gid": object, "Year": object}
        df = df.astype(convert_dict)
        # print(df.dtypes)
        GenINVtot = pd.DataFrame()
        GenINVtot = df.groupby(["Bus", "Tech", "Gid"])[
            "level"].sum().reset_index()
        GenINVtot = (
            GenINVtot[GenINVtot["level"] >= 0]
            .sort_values(by="level", ascending=False)
            .reset_index(drop=True)
        )
        convert_dict = {"Bus": int, "Gid": int}
        GenINVtot = GenINVtot.astype(convert_dict)
        GenINVtot = pd.concat(
            [GenINVtot, pd.DataFrame(
                columns=["Bus_Name", "Lon", "Lat", "rad"])]
        )

        for i in np.arange(0, np.shape(GenINVtot)[0]):
            GenINVtot_slice = GenINVtot.iloc[i]
            Bus = GenINVtot_slice["Bus"]
            lev = GenINVtot_slice["level"]

            GenINVtot.at[i, "Bus_Name"] = Circles.loc[Circles["Number"] == Bus][
                "Name"
            ].values[0]
            GenINVtot.at[i, "Lon"] = Circles.loc[Circles["Number"] == Bus][
                "Lon"
            ].values[0]
            GenINVtot.at[i, "Lat"] = Circles.loc[Circles["Number"] == Bus][
                "Lat"
            ].values[0]
            if lev > 0 and lev <= 1000:
                GenINVtot.at[i, "rad"] = 1
            elif lev > 1000 and lev <= 2000:
                GenINVtot.at[i, "rad"] = 2
            elif lev > 2000 and lev <= 3000:
                GenINVtot.at[i, "rad"] = 3
            elif lev > 3000 and lev <= 4000:
                GenINVtot.at[i, "rad"] = 4
            elif lev > 4000 and lev <= 5000:
                GenINVtot.at[i, "rad"] = 5
            elif lev > 5000 and lev <= 6000:
                GenINVtot.at[i, "rad"] = 6
            elif lev > 6000:
                GenINVtot.at[i, "rad"] = 7
        convert_dict = {
            "Bus_Name": str,
            "Tech": str,
            "Bus": int,
            "Gid": int,
            "Lon": float,
            "Lat": float,
            "rad": float,
        }
        GenINVtot = GenINVtot.astype(convert_dict)

        HMSolar = folium.FeatureGroup(name="Solar (orange)")
        HMWind = folium.FeatureGroup(name="Wind (green)")
        HMGas_CC = folium.FeatureGroup(name="Gas CC(red)")
        circrad = 10  # Summary.Parameter[3]/500
        colors = ["orange", " green", "red"]
        lat_dev = 0.5
        lon_dev = 0.2
        # lat_dev2 = 0.2
        # lon_dev2 = 0.6

        for i in np.arange(0, np.shape(GenINVtot)[0]):
            popup = (
                "Bus: "
                + str(round(GenINVtot.Bus[i], 0))
                + " Gen: "
                + str(GenINVtot.Gid[i])
                + " Tech: "
                + str(GenINVtot.Tech[i])
                + " Invested Cap: "
                + str(round(GenINVtot.level[i], 2))
                + " MW"
            )

            if GenINVtot.Tech[i] == "SolarNew":
                HMSolar.add_child(
                    folium.RegularPolygonMarker(
                        [GenINVtot.Lat[i], GenINVtot.Lon[i]],
                        number_of_sides=5,
                        radius=(circrad * GenINVtot.rad[i]),
                        color=colors[0],
                        fill=True,
                        fill_color=colors[0],
                        fill_opacity=0.8,
                        tooltip=popup,
                        popup=popup,
                    )
                )
                # -0.5, +0.5*lon_dev
            if GenINVtot.Tech[i] == "WindNew":
                HMWind.add_child(
                    folium.RegularPolygonMarker(
                        [GenINVtot.Lat[i], GenINVtot.Lon[i]],
                        number_of_sides=3,
                        radius=(circrad * GenINVtot.rad[i]),
                        color=colors[1],
                        fill=True,
                        fill_color=colors[1],
                        fill_opacity=0.5,
                        tooltip=popup,
                        popup=popup,
                    )
                )
                # +0.13, +lon_dev
            if GenINVtot.Tech[i] == "CCNew":
                HMGas_CC.add_child(
                    folium.CircleMarker(
                        [GenINVtot.Lat[i], GenINVtot.Lon[i]],
                        number_of_sides=3,
                        radius=(circrad * GenINVtot.rad[i]),
                        color=colors[2],
                        fill=True,
                        fill_color=colors[2],
                        fill_opacity=0.5,
                        tooltip=popup,
                        popup=popup,
                    )
                )
                # 0, +0.25*lon_dev

        miso_map.add_child(HMSolar)
        miso_map.add_child(HMWind)
        miso_map.add_child(HMGas_CC)

    ##################### Mapping Invested DR################################
    if DR_map != 0 and "pv_DRTotInv" in dfvars.keys():
        print("MAPPING INVESTED DEMAND RESPONSE")
        df = dfvars["pv_DRTotInv"].sort_values(by="level", ascending=False)
        df = df.loc[df["level"] > 0.001][["Bus", "Tech", "Year", "level"]]
        # print(df.dtypes)
        convert_dict = {"Bus": object, "Tech": object, "Year": object}
        df = df.astype(convert_dict)
        # print(df.dtypes)
        DRINVtot = pd.DataFrame()
        DRINVtot = df.groupby(["Bus", "Tech"])["level"].sum().reset_index()
        DRINVtot = (
            DRINVtot[DRINVtot["level"] >= 0]
            .sort_values(by="level", ascending=False)
            .reset_index(drop=True)
        )
        convert_dict = {"Bus": int}
        DRINVtot = DRINVtot.astype(convert_dict)
        DRINVtot = pd.concat(
            [DRINVtot, pd.DataFrame(columns=["Bus_Name", "Lon", "Lat", "rad"])]
        )

        for i in np.arange(0, np.shape(DRINVtot)[0]):
            DRINVtot_slice = DRINVtot.iloc[i]
            Bus = DRINVtot_slice["Bus"]
            lev = DRINVtot_slice["level"]

            DRINVtot.at[i, "Bus_Name"] = Circles.loc[Circles["Number"] == Bus][
                "Name"
            ].values[0]
            DRINVtot.at[i, "Lon"] = Circles.loc[Circles["Number"] == Bus]["Lon"].values[
                0
            ]
            DRINVtot.at[i, "Lat"] = Circles.loc[Circles["Number"] == Bus]["Lat"].values[
                0
            ]
            if lev > 0 and lev <= 1000:
                DRINVtot.at[i, "rad"] = 1
            elif lev > 1000 and lev <= 2000:
                DRINVtot.at[i, "rad"] = 2
            elif lev > 2000 and lev <= 3000:
                DRINVtot.at[i, "rad"] = 3
            elif lev > 3000 and lev <= 4000:
                DRINVtot.at[i, "rad"] = 4
            elif lev > 4000 and lev <= 5000:
                DRINVtot.at[i, "rad"] = 5
            elif lev > 5000 and lev <= 6000:
                DRINVtot.at[i, "rad"] = 6
            elif lev > 6000:
                DRINVtot.at[i, "rad"] = 7
        convert_dict = {
            "Bus_Name": str,
            "Tech": str,
            "Bus": int,
            "Lon": float,
            "Lat": float,
            "rad": float,
        }
        DRINVtot = DRINVtot.astype(convert_dict)
        HMDRHC = folium.FeatureGroup(name="DRHC (grey)")
        HMDREV = folium.FeatureGroup(name="DREV (purple)")
        HMDRBLight = folium.FeatureGroup(name="DRBlight (brown)")
        HMWTP = folium.FeatureGroup(name="WTP (blue)")
        HMDRRes = folium.FeatureGroup(name="DRRes (cyan)")
        circrad = 10  # Summary.Parameter[3]/500
        lat_dev = 0.5
        lon_dev = 0.2
        lat_dev2 = 0.2
        lon_dev2 = 0.6
        colors = ["grey", "purple", "brown", "blue", "cyan"]

        for i in np.arange(0, np.shape(DRINVtot)[0]):
            popup = (
                "Bus: "
                + str(round(DRINVtot.Bus[i], 0))
                + " Tech: "
                + str(DRINVtot.Tech[i])
                + " Invested Cap: "
                + str(round(DRINVtot.level[i], 2))
                + " MW"
            )

            if DRINVtot.Tech[i] == "DRHC":
                HMDRHC.add_child(
                    folium.RegularPolygonMarker(
                        [DRINVtot.Lat[i] - 0.5, DRINVtot.Lon[i] + 0.5 * lon_dev],
                        number_of_sides=6,
                        radius=(circrad * DRINVtot.rad[i]),
                        color=colors[0],
                        fill=True,
                        fill_color=colors[0],
                        fill_opacity=0.5,
                        tooltip=popup,
                        popup=popup,
                    )
                )
                # -0.5, +0.5*lon_dev
            if DRINVtot.Tech[i] == "DREV":
                HMDREV.add_child(
                    folium.RegularPolygonMarker(
                        [DRINVtot.Lat[i] + 0.13, DRINVtot.Lon[i] + lon_dev],
                        number_of_sides=5,
                        radius=(circrad * DRINVtot.rad[i]),
                        color=colors[1],
                        fill=True,
                        fill_color=colors[1],
                        fill_opacity=0.8,
                        tooltip=popup,
                        popup=popup,
                    )
                )
                # +0.13, +lon_dev
            if DRINVtot.Tech[i] == "DRBLDLight":
                HMDRBLight.add_child(
                    folium.RegularPolygonMarker(
                        [DRINVtot.Lat[i], DRINVtot.Lon[i] + 0.25 * lon_dev],
                        number_of_sides=4,
                        radius=(circrad * DRINVtot.rad[i]),
                        color=colors[2],
                        fill=True,
                        fill_color=colors[2],
                        fill_opacity=0.5,
                        tooltip=popup,
                        popup=popup,
                    )
                )

                # -0.5, +0.5*lon_dev
            if DRINVtot.Tech[i] == "WTP":
                HMWTP.add_child(
                    folium.RegularPolygonMarker(
                        [DRINVtot.Lat[i], DRINVtot.Lon[i] + 0.5 * lon_dev2],
                        number_of_sides=3,
                        radius=(circrad * DRINVtot.rad[i]),
                        color=colors[3],
                        fill=True,
                        fill_color=colors[3],
                        fill_opacity=0.5,
                        tooltip=popup,
                        popup=popup,
                    )
                )
                # +0.13, +lon_dev
            if DRINVtot.Tech[i] == "DRRes":
                HMDRRes.add_child(
                    folium.CircleMarker(
                        [DRINVtot.Lat[i], DRINVtot.Lon[i] + 0.25 * lon_dev2],
                        radius=(circrad * DRINVtot.rad[i]),
                        color=colors[4],
                        fill=True,
                        fill_color=colors[4],
                        fill_opacity=0.8,
                        tooltip=popup,
                        popup=popup,
                    )
                )

        miso_map.add_child(HMDRHC)
        miso_map.add_child(HMDREV)
        miso_map.add_child(HMDRBLight)
        miso_map.add_child(HMWTP)
        miso_map.add_child(HMDRRes)

    print("CREATING LEGEND AND SAVING MAP....")

    if "pv_DRTotInv" in dfvars.keys():
        # labels = ["Ln <1 GW","Ln 1<2 GW","Ln 2<3 GW","Ln 3<4 GW","Ln 4< GW", "Solar", "Wind", "GT CC", "DR HC", "DR EV", "DR Bldg Light", "DR WTP", "DR Residential"]
        # colors = ['black', 'palegreen','LimeGreen','Green','DarkGreen', 'orange', 'green', 'red', 'grey', 'purple', 'brown', 'blue', 'cyan']
        map_legend(miso_map, dfvars)
        map_name = "MISO Line Gen DR Investments"
    else:
        # labels = ["Ln <1 GW","Ln 1<2 GW","Ln 2<3 GW","Ln 3<4 GW","Ln <4 GW", "Solar", "Wind", "GT CC"]
        # colors = ['black', 'palegreen','LimeGreen','Green','DarkGreen', 'ornage', 'green', 'red']
        map_legend(miso_map, dfvars)
        map_name = "MISO Line Gen Investments"

    # miso_map.add_legend(title="Legend", labels = labels, colors = colors)
    miso_map.add_child(folium.map.LayerControl())
    miso_map.save(map_name + ".html")


def map_legend(miso_map, dfvars):
    if "pv_DRTotInv" in dfvars.keys():
        template = """
        {% macro html(this, kwargs) %}

        <!doctype html>
        <html lang="en">
        <head>
          <meta charset="utf-8">
          <meta name="viewport" content="width=device-width, initial-scale=1">
          <title>jQuery UI Draggable - Default functionality</title>
          <link rel="stylesheet" href="//code.jquery.com/ui/1.12.1/themes/base/jquery-ui.css">

          <script src="https://code.jquery.com/jquery-1.12.4.js"></script>
          <script src="https://code.jquery.com/ui/1.12.1/jquery-ui.js"></script>

          <script>
          $( function() {
            $( "#maplegend" ).draggable({
                            start: function (event, ui) {
                                $(this).css({
                                    right: "auto",
                                    top: "auto",
                                    bottom: "auto"
                                });
                            }
                        });
        });

          </script>
        </head>
        <body>


        <div id='maplegend' class='maplegend'
            style='position: absolute; z-index:9999; border:2px solid grey; background-color:rgba(255, 255, 255, 0.8);
             border-radius:6px; padding: 10px; font-size:14px; right: 20px; bottom: 20px;'>

        <div class='legend-title'>Investments (GW)</div>
        <div class='legend-scale'>
          <ul class='legend-labels'>
            <li><span style='background:black;opacity:0.7;'></span>Ln <1</li>
            <li><span style='background:palegreen;opacity:0.7;'></span>Ln 1<2</li>
            <li><span style='background:LimeGreen;opacity:0.7;'></span>Ln 2<3</li>
            <li><span style='background:green;opacity:0.7;'></span>Ln 3<4</li>
            <li><span style='background:DarkGreen;opacity:0.7;'></span>Ln 4<</li>
            <li><span style='background:orange;opacity:0.7;'></span>Solar</li>
            <li><span style='background:green;opacity:0.7;'></span>Wind</li>
            <li><span style='background:red;opacity:0.7;'></span>GT CC</li>
            <li><span style='background:grey;opacity:0.7;'></span>DR HC</li>
            <li><span style='background:purple;opacity:0.7;'></span>DR EV</li>
            <li><span style='background:brown;opacity:0.7;'></span>DR Bldg Light</li>
            <li><span style='background:blue;opacity:0.7;'></span>DR WTP</li>
            <li><span style='background:cyan;opacity:0.7;'></span>DR Residential</li>

          </ul>
        </div>
        </div>

        </body>
        </html>

        <style type='text/css'>
          .maplegend .legend-title {
            text-align: left;
            margin-bottom: 5px;
            font-weight: bold;
            font-size: 90%;
            }
          .maplegend .legend-scale ul {
            margin: 0;
            margin-bottom: 5px;
            padding: 0;
            float: left;
            list-style: none;
            }
          .maplegend .legend-scale ul li {
            font-size: 80%;
            list-style: none;
            margin-left: 0;
            line-height: 18px;
            margin-bottom: 2px;
            }
          .maplegend ul.legend-labels li span {
            display: block;
            float: left;
            height: 16px;
            width: 30px;
            margin-right: 5px;
            margin-left: 0;
            border: 1px solid #999;
            }
          .maplegend .legend-source {
            font-size: 80%;
            color: #777;
            clear: both;
            }
          .maplegend a {
            color: #777;
            }
        </style>
        {% endmacro %}"""

        macro = MacroElement()
        macro._template = Template(template)

        miso_map.get_root().add_child(macro)
    else:
        template = """
        {% macro html(this, kwargs) %}

        <!doctype html>
        <html lang="en">
        <head>
          <meta charset="utf-8">
          <meta name="viewport" content="width=device-width, initial-scale=1">
          <title>jQuery UI Draggable - Default functionality</title>
          <link rel="stylesheet" href="//code.jquery.com/ui/1.12.1/themes/base/jquery-ui.css">

          <script src="https://code.jquery.com/jquery-1.12.4.js"></script>
          <script src="https://code.jquery.com/ui/1.12.1/jquery-ui.js"></script>

          <script>
          $( function() {
            $( "#maplegend" ).draggable({
                            start: function (event, ui) {
                                $(this).css({
                                    right: "auto",
                                    top: "auto",
                                    bottom: "auto"
                                });
                            }
                        });
        });

          </script>
        </head>
        <body>


        <div id='maplegend' class='maplegend'
            style='position: absolute; z-index:9999; border:2px solid grey; background-color:rgba(255, 255, 255, 0.8);
             border-radius:6px; padding: 10px; font-size:14px; right: 20px; bottom: 20px;'>

        <div class='legend-title'>Investments (GW)</div>
        <div class='legend-scale'>
          <ul class='legend-labels'>
            <li><span style='background:black;opacity:0.7;'></span>Ln <1</li>
            <li><span style='background:palegreen;opacity:0.7;'></span>Ln 1<2</li>
            <li><span style='background:LimeGreen;opacity:0.7;'></span>Ln 2<3</li>
            <li><span style='background:green;opacity:0.7;'></span>Ln 3<4</li>
            <li><span style='background:DarkGreen;opacity:0.7;'></span>Ln 4< GW</li>
            <li><span style='background:orange;opacity:0.7;'></span>Solar</li>
            <li><span style='background:green;opacity:0.7;'></span>Wind</li>
            <li><span style='background:red;opacity:0.7;'></span>GT CC</li>
          </ul>
        </div>
        </div>

        </body>
        </html>

        <style type='text/css'>
          .maplegend .legend-title {
            text-align: left;
            margin-bottom: 5px;
            font-weight: bold;
            font-size: 90%;
            }
          .maplegend .legend-scale ul {
            margin: 0;
            margin-bottom: 5px;
            padding: 0;
            float: left;
            list-style: none;
            }
          .maplegend .legend-scale ul li {
            font-size: 80%;
            list-style: none;
            margin-left: 0;
            line-height: 18px;
            margin-bottom: 2px;
            }
          .maplegend ul.legend-labels li span {
            display: block;
            float: left;
            height: 16px;
            width: 30px;
            margin-right: 5px;
            margin-left: 0;
            border: 1px solid #999;
            }
          .maplegend .legend-source {
            font-size: 80%;
            color: #777;
            clear: both;
            }
          .maplegend a {
            color: #777;
            }
        </style>
        {% endmacro %}"""

        macro = MacroElement()
        macro._template = Template(template)

        miso_map.get_root().add_child(macro)


def energy_plot(pg, genres, drres):
    # calculating pg for conventional and renewable generations for summer #!3
    pgs3 = pg.loc[(pg["Season"] == "Summer") & (pg["Block"] == "3")]
    # conventional generators output per year
    pgs3ren = pgs3.loc[
        (pgs3["Tech"].isin(["Solar", "SolarNew", "Wind", "WindNew", "Hydro"]))
    ]
    pgs3con = pgs3.loc[
        ~(pgs3["Tech"].isin(["Solar", "SolarNew", "Wind", "WindNew", "Hydro"]))
    ]
    con_energy = pgs3con.groupby("Year")[["level"]].sum()["level"]
    ren_energy = pgs3ren.groupby("Year")[["level"]].sum()["level"]
    con_energy.columns = ["con"]

    rencon = pd.DataFrame()
    rencon["con"] = con_energy
    rencon["ren"] = ren_energy
    rencon.plot.line()

    # generation reserve
    genress3 = genres.loc[(genres["Season"] == "Summer")
                          & (genres["Block"] == "3")]
    # genress3ren = genress3.loc[(genress3['Tech'].isin(['Solar', 'SolarNew', 'Wind', 'WindNew']))]
    genress3con = genress3.loc[
        ~(genress3["Tech"].isin(["Solar", "SolarNew", "Wind", "WindNew"]))
    ]

    con_res = (
        genress3con.groupby(["Year", "ResType"])[["level"]]
        .sum()["level"]
        .unstack()
        .fillna(value=0)
    )
    # essentially hydro only
    # ren_res = genress3ren.groupby(['Year', 'ResType'])[['L']].sum()['L'].unstack().fillna(value=0)

    # DR
    drs3 = drres.loc[(drres["Season"] == "Summer") & (drres["Block"] == "3")]
    dr_res = (
        drs3.groupby(["Year", "ResType"])[["level"]]
        .sum()["level"]
        .unstack()
        .fillna(value=0)
    )

    for col in con_res.columns:
        resplot = pd.DataFrame()
        resplot["con"] = con_res[col]
        resplot["ren"] = dr_res[col]
        resplot.plot.line()

    import ipdb

    ipdb.set_trace()  # XXX
    # plt.show()


def costvars(dfvars, selvars):
    dfs = [
        {kk: dfvars[kk]["level"]}
        for kk in selvars
        if (kk.startswith(r"pv_C_") and kk in dfvars.keys())
    ]
    dfl = []
    for d in dfs:
        for k, v in d.items():
            if not v.empty:
                v.index = [k]
                dfl.append(v)
    df = pd.concat(dfl).reset_index()
    df.columns = ["Item", "Cost (M$)"]
    df = df.set_index(["Item"])
    df.loc["Total"] = dfvars["fv_OBJ_CEP"]["level"].values
    df["%"] = df["Cost (M$)"] / dfvars["fv_OBJ_CEP"]["level"].values * 100
    return df.reset_index()


def write_excel(dfvars, selvars, resfile="results.xlsx"):
    filePath = resfile
    if os.path.exists(filePath):
        os.remove(filePath)
    else:
        print("Can not delete the file as it doesn't exists")

    for var in selvars:
        append_df_to_excel(resfile, dfvars[var], sheet_name=var, index=False)
    return


def rename_cols(dfvars, selvars, GAMSCOLS):
    for dvars, cols in selvars.items():
        for var in dvars:
            if var in dfvars.keys():
                print("working on ", var)
                # removing the values below 0.00001
                dfvars[var] = dfvars[var].loc[dfvars[var]["level"] > 0.00001]
                dfvars[var].columns = cols + GAMSCOLS
                if "Bus" in dfvars[var].columns:
                    dfvars[var]["Bus"] = dfvars[var]["Bus"].astype(int)
                    dfvars[var] = dfvars[var].loc[dfvars[var]["Bus"] > 150]

    return dfvars


def loadshed(dfvars):
    if dfvars["pv_LoadShed"]["level"].sum() > 0:
        print("we are having load shedding!!!")
        print("load shedding; stop")
        print(dfvars["pv_LoadShed"]["level"])


def vars_only(dfs, varsonly=True):
    # take only variables; they start with pv_ fv_ nv_
    if varsonly:
        dfvars = {
            k: v.records
            for k, v in dfs.items()
            if k.lower().startswith(("pv_", "fv_", "nv_"))
        }
    else:
        dfvars = {k: v.records for k, v in dfs.items()}
    return dfvars


def plt_fuelyear(df2, var):
    """I set to show only the values greater than 0.001 since it has very small values"""
    df2 = df2.loc[df2["Bus"] > 145]
    df = df2.groupby(["Fuel", "Year"]).sum("level")[["level"]]
    df = df.loc[df["level"] > 0.001]
    if not df.empty:
        ax = df["level"].unstack().plot.bar(stacked=True, title=var)
        ax.set_ylabel("MW")
        # ax = df["level"].unstack().T.plot.bar(stacked=True, title=var)
        # ax.set_ylabel('MW')


def plt_techyear_firstlastyear(df2, var, years=["2020", "2038"]):
    """I set to show only the values greater than 0.001 since it has very small values"""
    df2 = df2.loc[df2["Bus"] > 145]
    df = df2.groupby(["Tech", "Year"]).sum("level")[["level"]]
    df = df.loc[df["level"] > 0.001]
    if not df.empty:
        ax = df["level"].unstack()[years].plot.bar(stacked=False, title=var)
        ax.set_ylabel("MW")
        # ax = df["level"].unstack().T.plot.bar(stacked=True, title=var)
        # ax.set_ylabel('MW')


def label_function(val):
    return f"{val / 100 * len(df):.0f}\n{val:.0f}%"


def plot_pie(df, var):
    print(var)
    fig, (ax1, ax2) = plt.subplots(
        ncols=2, figsize=(10, 5), constrained_layout=True)
    df2 = (
        df.groupby(["Tech", "Year"])
        .sum("level")["level"]
        .unstack()["2020"]
        .fillna(value=0)
    )

    import ipdb

    ipdb.set_trace()  # XXX
    df2.plot.pie(
        autopct=label_function,
        textprops={"fontsize": 20},
        colors=["tomato", "gold", "skyblue"],
        ax=ax1,
        subplots=True,
    )
    import ipdb

    ipdb.set_trace()  # XXX
    # df.groupby(["Tech", "Year"]).sum('L')[["level"]].plot.pie(autopct=label_function, textprops={'fontsize': 20},
    #                                  colors=['violet', 'lime'], ax=ax2)
    ax1.set_ylabel("Per country", size=22)
    ax2.set_ylabel("Per gender", size=22)
    plt.tight_layout()


def plt_techyear(df2, var):
    """I set to show only the values greater than 0.001 since it has very small values"""
    print(var)
    print(df2)
    df2 = df2.loc[df2["Bus"] > 145]
    df = df2.groupby(["Tech", "Year"]).sum("level")[["level"]]
    df = df.loc[df["level"] > 0.001]
    if not df.empty:
        ax = df["level"].unstack().plot.bar(stacked=True, title=var)
        ax.set_ylabel("MW")
        # ax = df["level"].unstack().T.plot.bar(stacked=True, title=var)
        # ax.set_ylabel('MW')


def plt_techyear_pie(df2, var):
    """I set to show only the values greater than 0.001 since it has very small values"""
    df = df2.groupby(["Tech", "Year"]).sum("level")[["level"]]

    print(var)
    df = df["level"].unstack().fillna(value=0)
    labels = df.columns
    # make the pie circular by setting the aspect ratio to 1
    import ipdb

    ipdb.set_trace()  # XXX

    ax = df["level"].unstack().plot.pie(subplots=True, autopct="%1.1f%%")
    for y in df["level"].unstack().columns:
        if not df.empty:
            ax = df["level"].unstack().plot.pie(
                y=y, subplots=False, autopct="%1.1f%%")


def plt_seasonblock_firstlastyear(df4, var, years=["2020", "2038"], name=""):
    """I set to show only the values greater than 0.001 since it has very small values"""
    for j in ["3"]:
        df3 = df4.loc[df4["Block"] == j]
        for i in ["Summer"]:
            df2 = df3.loc[df3["Season"] == i]
            df = df2.groupby(["Tech", "Year"]).sum("level")[["level"]]
            df = df.loc[df["level"] > 0.001]
            if not df.empty:
                dff = df["level"].unstack()
                dff.plot.bar(
                    stacked=False, title=name + " " + var + " " + i + " Block " + j
                )
                # df["level"].unstack().T.plot.bar(stacked=True,
                #                              title=var + ' ' + i + ' Block ' +
                #                              j)


def plt_seasonblock(df4, var):
    """I set to show only the values greater than 0.001 since it has very small values"""
    for j in ["1", "2", "3", "4"]:
        df3 = df4.loc[df4["Block"] == j]
        for i in ["Summer", "Fall", "Winter", "Spring", "Peak"]:
            df2 = df3.loc[df3["Season"] == i]
            df = df2.groupby(["Tech", "Year"]).sum("level")[["level"]]
            df = df.loc[df["level"] > 0.001]
            if not df.empty:
                df["level"].unstack().plot.bar(
                    stacked=True, title=var + " " + i + " Block " + j
                )
                # df["level"].unstack().T.plot.bar(stacked=True,
                #                              title=var + ' ' + i + ' Block ' +
                #                              j)


def read_gdx(ws, gdxfile):
    m = gt.Container(gdxfile, ws.system_directory)

    # df = rdgdx.to_dataframe( fields=["level", "M", "LO", "UP", "SCALE"])  # returns a dict of all dataframes
    # df = {k: rdgdx.data[k] for k in rdgdx.data.keys()}
    df = m.data

    return df


def writing_gdx(ws, data, gdxfile):
    """write a dictioanry of dataframes to a gdxfile"""

    # create empty gdx container
    gdx = gt.GdxContainer(ws.system_directory)
    gdx.rgdx()
    gdx.validate(data)
    std_data = gdx.standardize(data)
    gdx.add_to_gdx(std_data, standardize_data=True, inplace=True)
    gdx.write_gdx(gdxfile)

    return


def append_df_to_excel(
    filename,
    df,
    sheet_name="Sheet1",
    **to_excel_kwargs,
):
    """append to excel file"""

    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            **to_excel_kwargs,
        )
        return

    with pd.ExcelWriter(filename, mode="a", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, **to_excel_kwargs)

    return


def append_df_to_excel_dfdict(
    filename,
    dfdict,
    Iter=0,
    **to_excel_kwargs,
):
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            for i, k in enumerate(dfdict):
                df = dfdict[k]
                df.to_excel(
                    writer, sheet_name="Year" + str(k) + "_Iter" + str(Iter), index=False
                )
        return

    with pd.ExcelWriter(filename, mode="a", engine="openpyxl") as writer:
        for i, k in enumerate(dfdict):
            df = dfdict[k]
            df.to_excel(
                writer, sheet_name="Year" + str(k) + "_Iter" + str(Iter), index=False
            )

    return


def append_df_to_excel_old(
    filename,
    df,
    sheet_name="Sheet1",
    startrow=None,
    truncate_sheet=True,
    **to_excel_kwargs,
):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs,
        )
        return

    # ignore [engine] parameter if it was passed
    if "engine" in to_excel_kwargs:
        to_excel_kwargs.pop("engine")

    writer = pd.ExcelWriter(
        filename, engine="openpyxl", mode="a", if_sheet_exists="replace"
    )

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets.update({ws.title: ws for ws in writer.book.worksheets})

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def label_function(values):
    def my_autopct(pct):
        total = sum(values) / 1000
        val = int(round(pct * total / 100.0))
        # return "{p:.0f}%\n({v:d})".format(p=pct, v=val)
        # return "{p:.0f}% ({v:d})".format(p=pct, v=int(val))
        return "{p:.0f}%".format(p=pct, v=val)

    return my_autopct


def label_function_cap(values):
    def my_autopct(pct):
        total = sum(values) / 1000
        val = int(round(pct * total / 100.0))
        # return "{p:.0f}%\n({v:d})".format(p=pct, v=val)
        return "{p:.0f}%".format(p=pct)
        # return "{p:.0f}% ({v:d})".format(p=pct, v=val)

    return my_autopct


def get_colvars():
    GAMSCOLS = ["level", "marginal", "lower", "upper", "scale"]

    # commas are important people
    selvars = {
        ("pv_GenCapTot", "pv_GenInvYear", "pv_GenRetYear"): (
            [
                "Scen",
                "Bus",
                "Tech",
                "Gid",
                "Year",
            ]
        ),
        # generation dispatch
        (
            "pv_PG",
            "pv_GenCurt",
        ): (["Scen", "Bus", "Tech", "Gid", "Year", "Season", "Block"]),
        # reserves
        ("pv_GenRes",): (
            ["Scen", "ResType", "Bus", "Tech", "Gid", "Year", "Season", "Block"]
        ),
        ("pv_DRRes",): (["Scen", "ResType", "Bus", "Tech", "Year", "Season", "Block"]),
        ("pv_DRTotInv",): (["Scen", "Bus", "Tech", "Year"]),
        ("pv_ResReq",): (["Scen", "ResType", "Year"]),
        ("pv_ResChars",): (["Scen", "Char", "Year"]),
        # transmission
        (
            "fv_lineflow",
            "fv_resdnlineflow",
            "fv_resuplineflow",
            "nv_tielineflowDN",
            "pv_tielineflowUP",
        ): (["Scen", "From", "To", "Lid", "Year", "Season", "Block"]),
        (
            "pv_LineInv",
            "pv_ResUpLineInv",
        ): (["Scen", "From", "To", "Lid", "Year"]),
        ("pv_CarbonEmission",): (["Scen", "Year"]),
        ("pv_LoadShed",): (["Scen", "Bus", "Year", "Season", "Block"]),
        (
            "pv_C_GenInv",
            "pv_C_LineInv",
            "pv_C_DRResCost",
            "pv_C_DRResCosttot",
            "pv_C_FuelCost",
            "pv_C_GenFOM",
            "pv_C_GenVOM",
            "pv_C_GenResCost",
            "pv_C_loadshed",
            "pv_C_ResDRInv",
        ): ([]),
    }
    # XXX Other was added for capacity plots
    Techs = [
        "Biomass",
        "CC",
        "CCCCSNew",
        "CCNew",
        "CoalST",
        "DPVNew",
        "DRAgrPump",
        "DRBLDLight",
        "DRDataCent",
        "DREV",
        "DRHC",
        "DRManufC1",
        "DRManufC2",
        "DRManufC3",
        "DRRes",
        "EENew",
        "GasGT",
        "GasIC",
        "GasST",
        "GTNew",
        "Hydro",
        "Nuclear",
        "OilGT",
        "OilIC",
        "OilST",
        "PS",
        "Solar",
        "SolarNew",
        "STO",
        "STONew",
        "Waste",
        "Wind",
        "WindNew",
        "WTP",
        "WWTP",
        "Other",
    ]
    random.shuffle(Techs)

    return GAMSCOLS, selvars


def get_block(df, SN, BK):
    return df.loc[(df["Season"] == SN) & (df["Block"] == BK)]


def add_other(df, year="2020"):
    df["percentage"] = df[year] / df[year].sum()
    df5 = df.loc[df["percentage"] > 0.05]
    df4 = df.loc[df["percentage"] <= 0.05]
    df4["percentage"] = df4[year] / df4[year].sum()
    df = df5.append(pd.DataFrame(
        data={year: df4[year].sum()}, index=["Other"]))
    df["percentage"] = df[year] / df[year].sum()
    return df, df4


def pie_chart(df, dfrest, ax1, colours, year="2020"):
    """the dataframe should come in sorted"""
    angle = 360 * df["percentage"].to_list()[-1] / 2
    df = df.drop(columns="percentage")
    df.plot.pie(
        y=year,
        autopct=label_function(df[year].to_list()),
        textprops={"fontsize": 10},
        ax=ax1,
        labeldistance=1.04,
        shadow=False,
        startangle=angle,
        pctdistance=0.5,
        colors=[colours[key] for key in df.index],
        explode=[0.05] * df.shape[0],
        legend=None,
    )

    # bottom = .1
    # width = .1
    # # dfrest.plot.bar(y=year, ax=ax2, stacked=True)
    # # Adding from the top matches the legend.
    # for j, (height, label) in enumerate(reversed([*zip(dfrest["percentage"], dfrest.index)])):
    #     bottom -= height
    #     print(bottom, height, width)
    #     bc = ax2.bar(0, height, width, bottom=bottom, color='C0', label=label,
    #                  # alpha=0.1 + 0.25 * j
    #                  )
    #     # ax2.bar_label(bc, labels=[f"{height:.0%}"], label_type='center')
    #     ax2.bar_label(bc, labels=[dfrest.index.to_list()[j]], label_type='center')
    # ax2.set_title('Others')
    # ax2.legend()
    # ax2.axis('off')
    # ax2.set_xlim(- 2.5 * width, 2.5 * width)


def plot_energy(dfvars, colours, season, block):
    # ##############################################################################
    # ######################Energy Plots############################################
    # ##############################################################################

    myvar = "pv_PG"

    df = dfvars[myvar]
    fig, (ax1, ax2) = plt.subplots(
        ncols=2, figsize=(10, 5), constrained_layout=True, gridspec_kw={"wspace": 0.3}
    )
    fig.suptitle("Sources of Energy for {0}, block {1}".format(season, block))
    df = df.loc[(df["Season"] == season) & (df["Block"] == block)]
    year = "2020"
    if year in df["Year"].unique().to_list():
        df3 = (
            df.groupby(["Tech", "Year"])
            .sum("level")["level"]
            .unstack()[[year]]
            .fillna(value=0)
        )
        dfcap2020orig = df3  # why dfcap2020orig has "percentage" column?
        df3["percentage"] = df3[year] / df3[year].sum()
        # labels = list(df3.index) + ["Other"]
        df5 = df3.loc[df3["percentage"] > 0.05]
        df4 = df3.loc[df3["percentage"] <= 0.05]
        dfcap20 = df5.append(
            pd.DataFrame(data={year: df4[year].sum()}, index=["Other"])
        )
        dfcap20 = dfcap20.drop(columns="percentage")
        dfcap20 = dfcap20 * 1000
        dfcap20[year].plot.pie(
            autopct=label_function(dfcap20[year].tolist()),
            textprops={"fontsize": 10},
            ax=ax1,
            labeldistance=1.05,
            shadow=False,
            startangle=0,
            pctdistance=0.4,
            colors=[colours[key] for key in dfcap20.index],
            explode=[0.05] * dfcap20.shape[0],
        )
        print(dfcap20)
        ax1.set_ylabel("2020", size=12)
        ax1.yaxis.set_label_coords(-0.05, 0.5)
    year = "2038"
    if year in df["Year"].unique().to_list():
        df3 = (
            df.groupby(["Tech", "Year"])
            .sum("level")["level"]
            .unstack()[[year]]
            .fillna(value=0)
        )
        df3["percentage"] = df3[year] / df3[year].sum()
        df3 = df3.sort_index()
        dfcap2038orig = df3  # why dfcap2038orig has "percentage" column?
        df5 = df3.loc[df3["percentage"] > 0.02]
        df4 = df3.loc[df3["percentage"] <= 0.02]
        dfcap38 = df5.append(
            pd.DataFrame(data={year: df4[year].sum()}, index=["Other"])
        )
        dfcap38 = dfcap38.drop(columns="percentage")
        dfcap38 = dfcap38 * 1000
        dfcap38[year].plot.pie(
            autopct=label_function(dfcap38[year].tolist()),
            textprops={"fontsize": 10},
            ax=ax2,
            labeldistance=1.04,
            shadow=False,
            startangle=0,
            pctdistance=0.4,
            colors=[colours[key] for key in dfcap38.index],
            explode=[0.05] * dfcap38.shape[0],
        )
        print(dfcap38)
        ax2.set_ylabel("2038", size=12)
        ax2.yaxis.set_label_coords(-0.05, 0.5)
    plt.tight_layout()
    plt.savefig("./Figs/energy" + season + ".pdf")


def plot_resreq(dfvars):
    # ##############################################################################
    # ######################Reserve Requirements####################################
    # ##############################################################################
    fig, ax = plt.subplots(figsize=(10, 5), constrained_layout=True)
    resreq = dfvars["pv_ResReq"]
    resdfs = []
    ResTypes = ["RegUp", "RegDwn", "RampUp",
                "RampDwn", "STRUp", "STRDwn", "Cont"]
    for ResType in ResTypes:
        df = resreq.loc[
            # (resreq["Season"] == "Summer")
            # & (resreq["Block"] == "3")
            # & (resreq["ResType"] == ResType)
            (resreq["ResType"] == ResType)
        ].set_index("Year")
        df = df[["level"]]
        df.columns = [ResType]
        resdfs.append(df)
        df.plot(ax=ax)
        # df.plot()
    dfs = pd.concat(resdfs, axis=1)
    ax.grid(zorder=0)
    dfs.plot.bar(title="Reserve Requirements")
    ax.set_ylabel("MW", size=12)
    plt.tight_layout()
    plt.savefig("./Figs/Reserverequirements.pdf")


def reserve_plotsnew(dfvars, colours):
    ResTypes = ["RegUp", "RegDwn", "RampUp",
                "RampDwn", "STRUp", "STRDwn", "Cont"]
    DRRes = pd.DataFrame()
    GenRes = pd.DataFrame()
    if "pv_DRRes" in dfvars.keys():
        df = get_block(dfvars["pv_DRRes"], SN="Summer", BK="3")
        DRRes = df.groupby(["ResType", "Tech", "Year"])[
            "level"].sum().reset_index()
    if "pv_GenRes" in dfvars.keys():
        df2 = get_block(dfvars["pv_GenRes"], SN="Summer", BK="3")
        GenRes = df2.groupby(["ResType", "Tech", "Year"])[
            "level"].sum().reset_index()
    if not DRRes.empty and not GenRes.empty:
        df = (
            pd.merge(
                DRRes, GenRes, on=["ResType", "Tech", "Year"], how="outer"
            ).set_index(["ResType", "Tech", "Year"])
        ).fillna(0)
        df.columns = ["DR", "Gen"]
        df["tot"] = df["DR"] + df["Gen"]
        df.drop(columns=["DR", "Gen"], inplace=True)
    elif not GenRes.empty:
        GenRes.rename(columns={"level": "tot"}, inplace=True)
        df = GenRes.set_index(["ResType", "Tech", "Year"])
    elif not DRRes.empty:
        DRRes.rename(columns={"level": "tot"}, inplace=True)
        df = DRRes.set_index(["ResType", "Tech", "Year"])
    df = df["tot"].unstack(2).fillna(0)
    if "2038" in df.columns and "2020" in df.columns:
        gobj = df[["2020", "2038"]].groupby(
            df.index.get_level_values("ResType"))
    else:
        gobj = df.groupby(df.index.get_level_values("ResType"))

    for ResType, dff in gobj:
        fig, (ax1, ax2) = plt.subplots(
            ncols=2, figsize=(10, 5), constrained_layout=True
        )
        fig.suptitle(
            ResType + " Reserve (Summer, 14:00 -- 18:00) in 2020 and 2038")

        if "2020" in dff.columns:
            df20 = dff[["2020"]].sort_values(by="2020", ascending=False)
            df20 = df20[df20.any(axis=1)]
            df20.index = df20.index.droplevel(0)
            df20, df20rest = add_other(df20, year="2020")
            pie_chart(df20, df20rest, ax1, colours, year="2020")
            ax1.set_ylabel("2020", size=12)
            ax1.yaxis.set_label_coords(-0.05, 0.5)

        if "2038" in dff.columns:
            df38 = dff[["2038"]].sort_values(by="2038", ascending=False)
            df38 = df38[df38.any(axis=1)]
            df38.index = df38.index.droplevel(0)
            df38, df38rest = add_other(df38, year="2038")
            pie_chart(df38, df38rest, ax2, colours, year="2038")
            ax2.set_ylabel("2038", size=12)
            ax2.yaxis.set_label_coords(-0.05, 0.5)

        plt.tight_layout()
        plt.savefig("./Figs/" + ResType + ".pdf")


def GenInv_plots(dfvars, colours):
    invvars = ["pv_GenCapTot", "pv_GenInvYear", "pv_GenRetYear"]
    names = ["capacity", "geninvestments", "genretirements"]
    titles = ["Available Capacity", "Investments", "Retirements"]
    for inx, myvar in enumerate(invvars):
        print("plotting ", myvar)
        df = dfvars[myvar]
        plotbars(df, names[inx], titles[inx], colours)
        fig, (ax1, ax2) = plt.subplots(
            ncols=2, figsize=(10, 5), constrained_layout=True
        )
        fig.suptitle(titles[inx] + " in 2020 and 2038")
        year = "2020"
        if not df.loc[df["Year"] == year].empty:
            df3 = (
                df.groupby(["Tech", "Year"])
                .sum("level")["level"]
                .unstack()[[year]]
                .fillna(value=0)
            )
            dfcap2020orig = df3  # why dfcap2020orig has "percentage" column?
            df3["percentage"] = df3[year] / df3[year].sum()
            # labels = list(df3.index) + ["Other"]
            df5 = df3.loc[df3["percentage"] > 0.05]
            df4 = df3.loc[df3["percentage"] <= 0.05]
            dfcap20 = df5.append(
                pd.DataFrame(data={year: df4[year].sum()}, index=["Other"])
            )
            dfcap20 = dfcap20.drop(columns="percentage")
            dfcap20 = dfcap20 * 1000
            dfcap20[year].plot.pie(
                autopct=label_function_cap(dfcap20[year].tolist()),
                textprops={"fontsize": 10},
                ax=ax1,
                labeldistance=1.04,
                shadow=False,
                startangle=0,
                pctdistance=0.5,
                colors=[colours[key] for key in dfcap20.index],
                explode=[0.05] * dfcap20.shape[0],
            )
            ax1.set_ylabel("2020", size=12)
            ax1.yaxis.set_label_coords(-0.05, 0.5)
        year = "2038"
        df3 = []
        if not df.loc[df["Year"] == year].empty:
            df3 = (
                df.groupby(["Tech", "Year"])
                .sum("level")["level"]
                .unstack()[[year]]
                .fillna(value=0)
            )
            df3["percentage"] = df3[year] / df3[year].sum()
            df3 = df3.sort_index()
            dfcap2038orig = df3  # why dfcap2038orig has "percentage" column?
            df5 = df3.loc[df3["percentage"] > 0.05]
            df4 = df3.loc[df3["percentage"] <= 0.05]
            dfcap38 = df5.append(
                pd.DataFrame(data={year: df4[year].sum()}, index=["Other"])
            )
            dfcap38 = dfcap38.drop(columns="percentage")
            dfcap38 = dfcap38 * 1000
            dfcap38[year].plot.pie(
                autopct=label_function_cap(dfcap38[year].tolist()),
                textprops={"fontsize": 10},
                ax=ax2,
                labeldistance=1.1,
                shadow=False,
                startangle=0,
                pctdistance=0.5,
                colors=[colours[key] for key in dfcap38.index],
                explode=[0.05] * dfcap38.shape[0],
            )
            ax2.set_ylabel("2038", size=12)
            ax2.yaxis.set_label_coords(-0.05, 0.5)

        plt.tight_layout()
        plt.savefig("./Figs/" + names[inx] + ".pdf")


def plotbars(df, name, title, colours):
    dfcapbar = (
        df.groupby(["Tech", "Year"]).sum("level")[
            "level"].unstack().fillna(value=0)
    )
    if not dfcapbar.empty:
        fig, ax1 = plt.subplots(ncols=1, figsize=(
            10, 5), constrained_layout=True)
        fig.suptitle(title + " ")
        dfcapbar.plot.bar(
            stacked=True, ax=ax1, color=[colours[key] for key in dfcapbar.index]
        )
        ax1.set_ylabel("MW", size=12)
        ax1.yaxis.set_label_coords(-0.05, 0.5)
        plt.tight_layout()
        plt.savefig("./Figs/bar_tech_" + name + ".pdf")

        fig, ax1 = plt.subplots(ncols=1, figsize=(
            10, 5), constrained_layout=True)
        fig.suptitle(title + " ")
        dfcapbar.T.plot.bar(
            stacked=True, ax=ax1, color=[colours[key] for key in dfcapbar.index]
        )
        ax1.set_ylabel("MW", size=12)
        ax1.yaxis.set_label_coords(-0.05, 0.5)
        plt.legend(loc=(1.04, 0), ncol=2)
        plt.tight_layout()
        plt.savefig("./Figs/bar_Year_" + name + ".pdf")


def reserve_plots(dfvars, colours):
    ResTypes = ["RegUp", "RegDwn", "RampUp",
                "RampDwn", "STRUp", "STRDwn", "Cont"]
    # ResType = "RampUp"
    for ResType in ResTypes:
        df20 = []
        df38 = []
        gobj = dfvars["pv_DRRes"].groupby("ResType")
        for name, df in gobj:
            if name == ResType:
                print(name)
                year = "2038"
                df = df.loc[(df["Season"] == "Summer") & (df["Block"] == "3")]
                df4 = df.groupby(["Tech", "Year"]).sum(
                    "level")["level"].unstack()
                if year in df4.columns:
                    df3 = df4[[year]].fillna(value=0)
                    dfdr38 = df3.sort_index()
                    dfdr38 = dfdr38.loc[dfdr38["2038"] > 0]
                    df38.append(dfdr38)
                    print(year)
                    print(dfdr38)
                year = "2020"
                if year in df4.columns:
                    df2 = df4[[year]].fillna(value=0)
                    dfdr20 = df2.sort_index()
                    dfdr20 = dfdr20.loc[dfdr20["2020"] > 0]
                    df20.append(dfdr20)
                    print(year)
                    print(dfdr20)

        gobj = dfvars["pv_GenRes"].groupby("ResType")
        for name, df in gobj:
            if name == ResType:
                print(name)
                df = df.loc[(df["Season"] == "Summer") & (df["Block"] == "3")]
                df4 = df.groupby(["Tech", "Year"]).sum(
                    "level")["level"].unstack()
                year = "2020"
                if year in df4.columns:
                    dfg20 = df4[[year]].fillna(value=0)
                    dfg20 = dfg20.loc[dfg20["2020"] > 0]
                    df20.append(dfg20)
                    print(year)
                    print(dfg20)
                year = "2038"
                if year in df4.columns:
                    dfg38 = df4[[year]].fillna(value=0)
                    dfg38 = dfg38.loc[dfg38["2038"] > 0]
                    df38.append(dfg38)
                    print(year)
                    print(dfg38)

        # XXX are we converting to MW?
        if df20:
            year = "2020"
            df20 = pd.concat(df20) * 1000
            df20["percentage"] = df20[year] / df20[year].sum()
            df5 = df20.loc[df20["percentage"] > 0.01]
            df4 = df20.loc[df20["percentage"] <= 0.01]
            df20 = df5.append(
                pd.DataFrame(data={year: df4[year].sum()}, index=["Other"])
            )
            df20 = df20.drop(columns="percentage")

        if df38:
            year = "2038"
            df38 = pd.concat(df38) * 1000
            df38["percentage"] = df38["2038"] / df38["2038"].sum()
            df5 = df38.loc[df38["percentage"] > 0.01]
            df4 = df38.loc[df38["percentage"] <= 0.01]
            df38 = df5.append(
                pd.DataFrame(data={year: df4[year].sum()}, index=["Other"])
            )
            df38 = df38.drop(columns="percentage")

        fig, (ax1, ax2) = plt.subplots(
            ncols=2, figsize=(10, 5), constrained_layout=True
        )
        fig.suptitle(
            ResType + " Reserve (Summer, 14:00 -- 18:00) in 2020 and 2038")

        if not df20.empty:
            year = "2020"
            df20[year].plot.pie(
                autopct=label_function(df20[year].tolist()),
                textprops={"fontsize": 10},
                ax=ax1,
                labeldistance=1.04,
                shadow=False,
                startangle=0,
                pctdistance=0.5,
                colors=[colours[key] for key in df20.index],
                explode=[0.05] * df20[year].shape[0],
            )

        # df38 = pd.concat([df38, pd.DataFrame(data={"2038": 0.47812}, index=["Hydro"])])

        if not df38.empty:
            year = "2038"
            df38[year].plot.pie(
                autopct=label_function(df38[year].tolist()),
                textprops={"fontsize": 10},
                ax=ax2,
                labeldistance=1.04,
                shadow=False,
                rotatelabels=90,
                startangle=0,
                pctdistance=0.5,
                colors=[colours[key] for key in df38.index],
                explode=[0.05] * df38[year].shape[0],
            )

        ax1.set_ylabel("2020", size=12)
        ax1.yaxis.set_label_coords(-0.05, 0.5)
        ax2.set_ylabel("2038", size=12)
        ax2.yaxis.set_label_coords(-0.05, 0.5)
        plt.tight_layout()
        plt.savefig("./Figs/" + ResType + ".pdf")
